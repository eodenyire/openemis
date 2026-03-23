"""Student reports — profile list, performance, attendance, fees, health."""
import io
from datetime import date
from typing import Optional

import openpyxl
from sqlalchemy import func, case
from sqlalchemy.orm import Session

from app.reports.helpers import (
    style_header_row, style_data_rows, auto_width, add_title_row,
    excel_response, pdf_response, pdf_table, pdf_doc, pdf_header, get_styles,
    xl_header_fill, xl_bold, NAVY, RED, AMBER, GREEN, WHITE,
)
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.platypus import Spacer, Paragraph
from reportlab.lib.units import cm

from app.models.people import Student, StudentCourse
from app.models.core import AcademicYear, Course
from app.models.attendance import AttendanceLine, AttendanceSheet, AttendanceRegister
from app.models.fees import StudentFeeInvoice
from app.models.cbc import ReportCard, ReportCardLine, CBCGradeLevel
from app.models.health import StudentHealth, ClinicVisit


def _year(db, academic_year_id):
    return (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())


def _att_rate(db, student_id, year_id):
    total = (db.query(AttendanceLine)
             .join(AttendanceSheet, AttendanceSheet.id == AttendanceLine.sheet_id)
             .join(AttendanceRegister, AttendanceRegister.id == AttendanceSheet.register_id)
             .filter(AttendanceLine.student_id == student_id,
                     AttendanceRegister.academic_year_id == year_id).count())
    if not total:
        return None
    present = (db.query(AttendanceLine)
               .join(AttendanceSheet, AttendanceSheet.id == AttendanceLine.sheet_id)
               .join(AttendanceRegister, AttendanceRegister.id == AttendanceSheet.register_id)
               .filter(AttendanceLine.student_id == student_id,
                       AttendanceRegister.academic_year_id == year_id,
                       AttendanceLine.status == "present").count())
    return round(present / total * 100, 1)


# ── Student Master List ───────────────────────────────────────────────────────

def student_list_excel(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    q = db.query(Student, Course.name.label("grade"))\
        .join(StudentCourse, StudentCourse.student_id == Student.id)\
        .join(Course, Course.id == StudentCourse.course_id)\
        .filter(StudentCourse.active == True, Student.active == True)
    if year:
        q = q.filter(StudentCourse.academic_year_id == year.id)
    rows = q.order_by(Course.name, Student.last_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student List"
    headers = ["#", "Adm No", "NEMIS UPI", "First Name", "Last Name",
               "Gender", "DOB", "Grade", "Phone", "Email", "Status"]
    start = add_title_row(ws, f"Student Master List — {year.name if year else 'All Years'}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))
    for i, (s, grade) in enumerate(rows, 1):
        ws.append([i, s.admission_number, s.nemis_upi or "",
                   s.first_name, s.last_name,
                   str(s.gender) if s.gender else "",
                   str(s.date_of_birth) if s.date_of_birth else "",
                   grade, s.mobile or s.phone or "", s.email or "",
                   "Active" if s.active else "Inactive"])
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"students_{year.name if year else 'all'}.xlsx")


def student_list_pdf(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    q = db.query(Student, Course.name.label("grade"))\
        .join(StudentCourse, StudentCourse.student_id == Student.id)\
        .join(Course, Course.id == StudentCourse.course_id)\
        .filter(StudentCourse.active == True, Student.active == True)
    if year:
        q = q.filter(StudentCourse.academic_year_id == year.id)
    rows = q.order_by(Course.name, Student.last_name).all()

    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, f"Student Master List — {year.name if year else 'All'}",
                       f"Total: {len(rows)} students  |  Generated: {date.today()}")
    data = [["#", "Adm No", "NEMIS UPI", "Name", "Gender", "Grade", "DOB"]]
    for i, (s, grade) in enumerate(rows, 1):
        data.append([i, s.admission_number, s.nemis_upi or "—",
                     f"{s.first_name} {s.last_name}",
                     str(s.gender) if s.gender else "—",
                     grade,
                     str(s.date_of_birth) if s.date_of_birth else "—"])
    elems.append(pdf_table(data, [1*cm, 3*cm, 3.5*cm, 5*cm, 2*cm, 3*cm, 3*cm]))
    doc.build(elems)
    return pdf_response(buf, f"students_{year.name if year else 'all'}.pdf")


# ── Student Performance Report ────────────────────────────────────────────────

def student_performance_excel(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    students = db.query(Student).filter_by(active=True).all()
    level_weight = {"EE": 4, "ME": 3, "AE": 2, "BE": 1}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance"
    headers = ["Adm No", "Name", "Grade", "EE", "ME", "AE", "BE", "Total Subjects", "Avg Score", "Overall Level"]
    start = add_title_row(ws, f"Student Performance Report — {year.name if year else ''}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    fills = {"EE": PatternFill("solid", fgColor="CCFFCC"),
             "ME": PatternFill("solid", fgColor="D6E4F0"),
             "AE": PatternFill("solid", fgColor="FFF2CC"),
             "BE": PatternFill("solid", fgColor="FFCCCC")}

    for s in students:
        # Get current grade
        sc = db.query(StudentCourse).filter_by(student_id=s.id, active=True).first()
        grade = sc.course.name if sc and sc.course else "—"

        lines = (db.query(ReportCardLine)
                 .join(ReportCard, ReportCard.id == ReportCardLine.report_card_id)
                 .filter(ReportCard.student_id == s.id))
        if year:
            lines = lines.filter(ReportCard.academic_year_id == year.id)
        lines = lines.all()

        dist = {"EE": 0, "ME": 0, "AE": 0, "BE": 0}
        for l in lines:
            k = str(l.performance_level)
            dist[k] = dist.get(k, 0) + 1
        total = sum(dist.values()) or 1
        weighted = sum(level_weight.get(k, 0) * v for k, v in dist.items())
        avg = round(weighted / total, 2)
        overall = max(dist, key=dist.get) if lines else "—"

        row = [s.admission_number, f"{s.first_name} {s.last_name}", grade,
               dist["EE"], dist["ME"], dist["AE"], dist["BE"],
               len(lines), avg, overall]
        ws.append(row)
        if overall in fills:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fills[overall]

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"performance_{year.name if year else 'all'}.xlsx")


def student_performance_pdf(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    students = db.query(Student).filter_by(active=True).all()
    level_weight = {"EE": 4, "ME": 3, "AE": 2, "BE": 1}

    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, f"Student Performance Report — {year.name if year else ''}",
                       f"Generated: {date.today()}")
    data = [["Adm No", "Name", "Grade", "EE", "ME", "AE", "BE", "Avg", "Level"]]
    for s in students:
        sc = db.query(StudentCourse).filter_by(student_id=s.id, active=True).first()
        grade = sc.course.name if sc and sc.course else "—"
        lines = (db.query(ReportCardLine)
                 .join(ReportCard, ReportCard.id == ReportCardLine.report_card_id)
                 .filter(ReportCard.student_id == s.id))
        if year:
            lines = lines.filter(ReportCard.academic_year_id == year.id)
        lines = lines.all()
        dist = {"EE": 0, "ME": 0, "AE": 0, "BE": 0}
        for l in lines:
            dist[str(l.performance_level)] = dist.get(str(l.performance_level), 0) + 1
        total = sum(dist.values()) or 1
        weighted = sum(level_weight.get(k, 0) * v for k, v in dist.items())
        avg = round(weighted / total, 2)
        overall = max(dist, key=dist.get) if lines else "—"
        data.append([s.admission_number, f"{s.first_name} {s.last_name}", grade,
                     dist["EE"], dist["ME"], dist["AE"], dist["BE"], avg, overall])
    elems.append(pdf_table(data, [3*cm, 5.5*cm, 3*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2*cm, 2*cm]))
    doc.build(elems)
    return pdf_response(buf, f"performance_{year.name if year else 'all'}.pdf")


# ── Student Attendance Report ─────────────────────────────────────────────────

def student_attendance_excel(db: Session, academic_year_id: Optional[int] = None,
                              threshold: float = 75.0):
    year = _year(db, academic_year_id)
    q = db.query(
        Student.admission_number, Student.first_name, Student.last_name,
        func.count(AttendanceLine.id).label("total"),
        func.sum(case((AttendanceLine.status == "present", 1), else_=0)).label("present"),
        func.sum(case((AttendanceLine.status == "absent", 1), else_=0)).label("absent"),
        func.sum(case((AttendanceLine.status == "late", 1), else_=0)).label("late"),
    ).join(AttendanceLine, AttendanceLine.student_id == Student.id)\
     .join(AttendanceSheet, AttendanceSheet.id == AttendanceLine.sheet_id)\
     .join(AttendanceRegister, AttendanceRegister.id == AttendanceSheet.register_id)
    if year:
        q = q.filter(AttendanceRegister.academic_year_id == year.id)
    rows = q.group_by(Student.admission_number, Student.first_name, Student.last_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["Adm No", "First Name", "Last Name", "Present", "Absent", "Late", "Total", "Rate %", "Flag"]
    start = add_title_row(ws, f"Student Attendance Report — {year.name if year else ''}",
                          len(headers), f"Threshold: {threshold}%  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for r in rows:
        rate = round((r.present or 0) / (r.total or 1) * 100, 1)
        flag = "⚠ AT RISK" if rate < threshold else "✓ OK"
        ws.append([r.admission_number, r.first_name, r.last_name,
                   r.present or 0, r.absent or 0, r.late or 0,
                   r.total or 0, rate, flag])
        if rate < threshold:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"attendance_{year.name if year else 'all'}.xlsx")


# ── Student Health Report ─────────────────────────────────────────────────────

def student_health_excel(db: Session):
    records = db.query(StudentHealth).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Health Records"
    headers = ["Adm No", "Name", "Blood Group", "Height(cm)", "Weight(kg)", "BMI",
               "Allergies", "Conditions", "Emergency Contact", "Insurance", "Last Checkup"]
    start = add_title_row(ws, "Student Health Records", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for r in records:
        s = r.student
        conditions = ", ".join(c.name for c in r.conditions) if r.conditions else "None"
        ws.append([
            s.admission_number if s else "—",
            f"{s.first_name} {s.last_name}" if s else "—",
            r.blood_group or "—", r.height or "—", r.weight or "—",
            r.bmi or "—", r.allergies or "None", conditions,
            f"{r.emergency_contact_name} ({r.emergency_contact_relation})" if r.emergency_contact_name else "—",
            r.insurance_provider or "—",
            str(r.last_checkup_date) if r.last_checkup_date else "—",
        ])
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "student_health.xlsx")


def student_health_pdf(db: Session):
    records = db.query(StudentHealth).all()
    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, "Student Health Records",
                       f"Total records: {len(records)}  |  Generated: {date.today()}")
    data = [["Adm No", "Name", "Blood Grp", "Height", "Weight", "BMI", "Allergies", "Conditions"]]
    for r in records:
        s = r.student
        conditions = ", ".join(c.name for c in r.conditions[:2]) if r.conditions else "None"
        data.append([
            s.admission_number if s else "—",
            f"{s.first_name} {s.last_name}" if s else "—",
            r.blood_group or "—", r.height or "—", r.weight or "—", r.bmi or "—",
            (r.allergies or "None")[:20], conditions[:25],
        ])
    elems.append(pdf_table(data, [2.5*cm, 5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 3.5*cm, 4*cm]))
    doc.build(elems)
    return pdf_response(buf, "student_health.pdf")
