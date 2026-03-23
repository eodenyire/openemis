"""Shared helpers for Excel and PDF report generation."""
import io
from datetime import date
from typing import List, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY   = "1F4E79"
BLUE   = "2E75B6"
LIGHT  = "D6E4F0"
RED    = "FF4444"
AMBER  = "FFAA00"
GREEN  = "70AD47"
WHITE  = "FFFFFF"

# ── Excel helpers ─────────────────────────────────────────────────────────────

def xl_header_fill(color=NAVY):
    return PatternFill("solid", fgColor=color)

def xl_alt_fill():
    return PatternFill("solid", fgColor=LIGHT)

def xl_header_font():
    return Font(color=WHITE, bold=True, size=10)

def xl_bold():
    return Font(bold=True)

def xl_center():
    return Alignment(horizontal="center", vertical="center")

def xl_thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header_row(ws, row: int, ncols: int, color=NAVY):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = xl_header_fill(color)
        cell.font = xl_header_font()
        cell.alignment = xl_center()
        cell.border = xl_thin_border()


def style_data_rows(ws, start_row: int, ncols: int):
    for row in ws.iter_rows(min_row=start_row, max_col=ncols):
        for i, cell in enumerate(row):
            cell.border = xl_thin_border()
            if cell.row % 2 == 0:
                cell.fill = xl_alt_fill()


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


def add_title_row(ws, title: str, ncols: int, subtitle: str = ""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14, color=NAVY)
    cell.alignment = xl_center()
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font = Font(italic=True, size=10, color="555555")
        sub.alignment = xl_center()
        return 3  # next data row
    return 2


def excel_response(wb, filename: str):
    from fastapi.responses import StreamingResponse
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def pdf_response(buf: io.BytesIO, filename: str):
    from fastapi.responses import StreamingResponse
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── PDF helpers ───────────────────────────────────────────────────────────────

def get_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("ReportTitle", parent=styles["Title"],
                               fontSize=16, textColor=colors.HexColor(f"#{NAVY}"),
                               spaceAfter=4))
    styles.add(ParagraphStyle("SubTitle", parent=styles["Normal"],
                               fontSize=10, textColor=colors.grey, spaceAfter=8))
    styles.add(ParagraphStyle("SectionHead", parent=styles["Heading2"],
                               fontSize=11, textColor=colors.HexColor(f"#{BLUE}"),
                               spaceBefore=12, spaceAfter=4))
    return styles


def pdf_table(data: List[List[Any]], col_widths: List[float] = None) -> Table:
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{LIGHT}")]),
        ("FONTSIZE",      (0, 1), (-1, -1), 7.5),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
    ]))
    return t


def pdf_doc(buf: io.BytesIO, landscape_mode=False):
    ps = landscape(A4) if landscape_mode else A4
    return SimpleDocTemplate(buf, pagesize=ps,
                              leftMargin=1.5*cm, rightMargin=1.5*cm,
                              topMargin=1.5*cm, bottomMargin=1.5*cm)


def pdf_header(styles, title: str, subtitle: str = "") -> list:
    elems = [Paragraph(title, styles["ReportTitle"])]
    if subtitle:
        elems.append(Paragraph(subtitle, styles["SubTitle"]))
    elems.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor(f"#{NAVY}"), spaceAfter=8))
    return elems
