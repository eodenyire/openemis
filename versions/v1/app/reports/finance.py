"""Finance reports — fee collection, outstanding balances, payment methods."""
import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.reports.helpers import (
    style_header_row, style_data_rows, auto_width, add_title_row,
    excel_response, pdf_response, pdf_table, pdf_doc, pdf_header, get_styles,
)
from reportlab.platypus import Spacer, Paragraph
from reportlab.lib.units import cm

from app.models.people import Student
from app.models.core import AcademicYear, Course
from app.models.fees import StudentFeeInvoice, FeePayment, PaymentState


def _year(db, academic_year_id):
    return (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())


def fee_collection_excel(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    q = db.query(
        Student.admission_number, Student.first_name, Student.last_name,
        Course.name.label("grade"),
        StudentFeeInvoice.total_amount, StudentFeeInvoice.paid_amount,
        StudentFeeInvoice.state, StudentFeeInvoice.due_date,
    ).join(StudentFeeInvoice, StudentFeeInvoice.student_id == Student.id)\
     .join(Course, Course.id == StudentFeeInvoice.course_id, isouter=True)\
     .filter(Student.active == True)
    if year:
        q = q.filter(StudentFeeInvoice.academic_year_id == year.id)
    rows = q.order_by(Course.name, Student.last_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee Collection"
    headers = ["Adm No", "First Name", "Last Name", "Grade",
               "Invoiced (KES)", "Paid (KES)", "Balance (KES)", "Rate %", "Status", "Due Date"]
    start = add_title_row(ws, f"Fee Collection Report — {year.name if year else ''}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    fills = {
        "paid":     PatternFill("solid", fgColor="CCFFCC"),
        "partial":  PatternFill("solid", fgColor="FFF2CC"),
        "pending":  PatternFill("solid", fgColor="FFE0CC"),
        "overdue":  PatternFill("solid", fgColor="FFCCCC"),
    }
    totals = {"inv": 0, "paid": 0}
    for r in rows:
        inv = r.total_amount or 0
        paid = r.paid_amount or 0
        bal = inv - paid
        rate = round(paid / inv * 100, 1) if inv else 0
        totals["inv"] += inv
        totals["paid"] += paid
        ws.append([r.admission_number, r.first_name, r.last_name, r.grade or "—",
                   inv, paid, round(bal, 2), rate, str(r.state),
                   str(r.due_date) if r.due_date else "—"])
        fill = fills.get(str(r.state))
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    # Summary row
    ws.append(["", "", "", "TOTAL",
               round(totals["inv"], 2), round(totals["paid"], 2),
               round(totals["inv"] - totals["paid"], 2),
               round(totals["paid"] / totals["inv"] * 100, 1) if totals["inv"] else 0,
               "", ""])
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"fees_{year.name if year else 'all'}.xlsx")


def fee_collection_pdf(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    # Grade-level summary for PDF
    q = db.query(
        Course.name,
        func.count(StudentFeeInvoice.id).label("students"),
        func.sum(StudentFeeInvoice.total_amount).label("invoiced"),
        func.sum(StudentFeeInvoice.paid_amount).label("paid"),
    ).join(StudentFeeInvoice, StudentFeeInvoice.course_id == Course.id, isouter=True)
    if year:
        q = q.filter(StudentFeeInvoice.academic_year_id == year.id)
    rows = q.group_by(Course.name).order_by(Course.name).all()

    total_inv = sum(float(r.invoiced or 0) for r in rows)
    total_paid = sum(float(r.paid or 0) for r in rows)

    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf)
    elems = pdf_header(styles, f"Fee Collection Summary — {year.name if year else ''}",
                       f"Total Invoiced: KES {total_inv:,.2f}  |  "
                       f"Collected: KES {total_paid:,.2f}  |  "
                       f"Outstanding: KES {total_inv - total_paid:,.2f}  |  "
                       f"Generated: {date.today()}")
    data = [["Grade", "Students", "Invoiced (KES)", "Paid (KES)", "Balance (KES)", "Rate %"]]
    for r in rows:
        inv = float(r.invoiced or 0)
        paid = float(r.paid or 0)
        data.append([r.name, r.students,
                     f"{inv:,.0f}", f"{paid:,.0f}",
                     f"{inv - paid:,.0f}",
                     f"{round(paid / inv * 100, 1) if inv else 0}%"])
    data.append(["TOTAL", "", f"{total_inv:,.0f}", f"{total_paid:,.0f}",
                 f"{total_inv - total_paid:,.0f}",
                 f"{round(total_paid / total_inv * 100, 1) if total_inv else 0}%"])
    elems.append(pdf_table(data, [4*cm, 2.5*cm, 4*cm, 4*cm, 4*cm, 2.5*cm]))
    doc.build(elems)
    return pdf_response(buf, f"fees_{year.name if year else 'all'}.pdf")


def outstanding_fees_excel(db: Session, academic_year_id: Optional[int] = None):
    """Students with outstanding balances only."""
    year = _year(db, academic_year_id)
    q = db.query(
        Student.admission_number, Student.first_name, Student.last_name,
        Course.name.label("grade"),
        StudentFeeInvoice.total_amount, StudentFeeInvoice.paid_amount,
        StudentFeeInvoice.state,
    ).join(StudentFeeInvoice, StudentFeeInvoice.student_id == Student.id)\
     .join(Course, Course.id == StudentFeeInvoice.course_id, isouter=True)\
     .filter(Student.active == True,
             StudentFeeInvoice.state.in_(["pending", "partial", "overdue"]))
    if year:
        q = q.filter(StudentFeeInvoice.academic_year_id == year.id)
    rows = q.order_by(Course.name, Student.last_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding Fees"
    headers = ["Adm No", "Name", "Grade", "Invoiced", "Paid", "Balance", "Status"]
    start = add_title_row(ws, "Outstanding Fee Balances", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for r in rows:
        bal = (r.total_amount or 0) - (r.paid_amount or 0)
        ws.append([r.admission_number, f"{r.first_name} {r.last_name}",
                   r.grade or "—", r.total_amount or 0, r.paid_amount or 0,
                   round(bal, 2), str(r.state)])
        if str(r.state) == "overdue":
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"outstanding_fees_{year.name if year else 'all'}.xlsx")


def payment_methods_excel(db: Session, academic_year_id: Optional[int] = None):
    """Payment method breakdown."""
    year = _year(db, academic_year_id)
    q = db.query(
        FeePayment.payment_method,
        func.count(FeePayment.id).label("count"),
        func.sum(FeePayment.amount).label("total"),
    )
    if year:
        q = q.join(StudentFeeInvoice, StudentFeeInvoice.id == FeePayment.invoice_id)\
             .filter(StudentFeeInvoice.academic_year_id == year.id)
    rows = q.group_by(FeePayment.payment_method).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Methods"
    headers = ["Payment Method", "Transactions", "Total (KES)", "Share %"]
    start = add_title_row(ws, "Payment Method Breakdown", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    grand_total = sum(float(r.total or 0) for r in rows)
    for r in rows:
        total = float(r.total or 0)
        ws.append([str(r.payment_method) if r.payment_method else "Unknown",
                   r.count, round(total, 2),
                   f"{round(total / grand_total * 100, 1) if grand_total else 0}%"])
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "payment_methods.xlsx")
