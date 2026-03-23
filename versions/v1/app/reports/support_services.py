"""Support service reports — Library, Hostel, Transport, Health."""
import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill
from sqlalchemy.orm import Session

from app.reports.helpers import (
    style_header_row, style_data_rows, auto_width, add_title_row,
    excel_response, pdf_response, pdf_table, pdf_doc, pdf_header, get_styles,
)
from reportlab.platypus import Spacer
from reportlab.lib.units import cm

from app.models.library import Media, MediaMovement, MediaMovementState
from app.models.hostel import HostelBlock, HostelRoom, HostelAllocation, HostelAllocationState
from app.models.transportation import Vehicle, TransportRoute, TransportRouteStop, StudentTransport
from app.models.health import StudentHealth, ClinicVisit, VaccinationRecord, MedicalCondition


# ══════════════════════════════════════════════════════════════════════════════
# LIBRARY
# ══════════════════════════════════════════════════════════════════════════════

def library_catalogue_excel(db: Session):
    books = db.query(Media).filter_by(active=True).order_by(Media.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Library Catalogue"
    headers = ["#", "Title", "ISBN", "Internal Code", "Type", "Grade Level",
               "Subject Area", "Total Copies", "Available", "Issued"]
    start = add_title_row(ws, "Library Catalogue", len(headers),
                          f"Total titles: {len(books)}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for i, b in enumerate(books, 1):
        issued = b.total_copies - b.available_copies
        ws.append([i, b.name, b.isbn or "—", b.internal_code or "—",
                   b.media_type.name if b.media_type else "—",
                   b.grade_level or "—", b.subject_area or "—",
                   b.total_copies, b.available_copies, issued])
        if b.available_copies == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "library_catalogue.xlsx")


def library_movements_excel(db: Session):
    movements = db.query(MediaMovement).order_by(MediaMovement.issue_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Borrowing Records"
    headers = ["#", "Book Title", "Student", "Adm No", "Issue Date",
               "Due Date", "Return Date", "Status", "Fine (KES)"]
    start = add_title_row(ws, "Library Borrowing Records", len(headers),
                          f"Total records: {len(movements)}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    fills = {
        "overdue": PatternFill("solid", fgColor="FFCCCC"),
        "lost":    PatternFill("solid", fgColor="FF9999"),
        "issued":  PatternFill("solid", fgColor="FFF2CC"),
    }
    for i, m in enumerate(movements, 1):
        student = m.student
        ws.append([i, m.media.name if m.media else "—",
                   f"{student.first_name} {student.last_name}" if student else "—",
                   student.admission_number if student else "—",
                   str(m.issue_date), str(m.due_date) if m.due_date else "—",
                   str(m.return_date) if m.return_date else "—",
                   str(m.state), m.fine or 0])
        fill = fills.get(str(m.state))
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "library_movements.xlsx")


def library_overdue_pdf(db: Session):
    overdue = db.query(MediaMovement).filter(
        MediaMovement.state.in_(["overdue", "issued"]),
        MediaMovement.due_date < date.today()
    ).all()

    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    total_fines = sum(m.fine or 0 for m in overdue)
    elems = pdf_header(styles, "Overdue Books Report",
                       f"Overdue: {len(overdue)}  |  Total Fines: KES {total_fines}  |  "
                       f"Generated: {date.today()}")
    data = [["Book Title", "Student", "Adm No", "Due Date", "Days Overdue", "Fine (KES)"]]
    for m in overdue:
        student = m.student
        days = (date.today() - m.due_date).days if m.due_date else 0
        data.append([
            (m.media.name if m.media else "—")[:40],
            f"{student.first_name} {student.last_name}" if student else "—",
            student.admission_number if student else "—",
            str(m.due_date), days, m.fine or 0,
        ])
    elems.append(pdf_table(data, [6*cm, 4.5*cm, 3*cm, 3*cm, 3*cm, 3*cm]))
    doc.build(elems)
    return pdf_response(buf, "library_overdue.pdf")


# ══════════════════════════════════════════════════════════════════════════════
# HOSTEL
# ══════════════════════════════════════════════════════════════════════════════

def hostel_occupancy_excel(db: Session):
    blocks = db.query(HostelBlock).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hostel Occupancy"
    headers = ["Block", "Room", "Capacity", "Occupied", "Available", "State", "Monthly Fee (KES)"]
    start = add_title_row(ws, "Hostel Occupancy Report", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for block in blocks:
        rooms = db.query(HostelRoom).filter_by(block_id=block.id).all()
        for room in rooms:
            occupied = db.query(HostelAllocation).filter_by(
                room_id=room.id, state=HostelAllocationState.CONFIRMED).count()
            fee = room.room_type.monthly_fee if room.room_type else 0
            ws.append([block.name, room.name, room.capacity, occupied,
                       room.capacity - occupied, room.state, fee])
            if room.state == "maintenance":
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="EEEEEE")
            elif occupied >= room.capacity:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "hostel_occupancy.xlsx")


def hostel_allocations_excel(db: Session):
    allocations = db.query(HostelAllocation)\
        .filter_by(state=HostelAllocationState.CONFIRMED).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hostel Allocations"
    headers = ["#", "Student", "Adm No", "Block", "Room", "Check-In", "Monthly Fee (KES)"]
    start = add_title_row(ws, "Hostel Allocations — Current Boarders", len(headers),
                          f"Total boarders: {len(allocations)}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for i, a in enumerate(allocations, 1):
        student = a.student
        room = a.room
        block = room.block if room else None
        ws.append([i,
                   f"{student.first_name} {student.last_name}" if student else "—",
                   student.admission_number if student else "—",
                   block.name if block else "—",
                   room.name if room else "—",
                   str(a.check_in_date), a.monthly_fee or 0])
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "hostel_allocations.xlsx")


def hostel_allocations_pdf(db: Session):
    allocations = db.query(HostelAllocation)\
        .filter_by(state=HostelAllocationState.CONFIRMED).all()
    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, "Hostel Allocations — Current Boarders",
                       f"Total: {len(allocations)} boarders  |  Generated: {date.today()}")
    data = [["#", "Student", "Adm No", "Block", "Room", "Check-In", "Fee/Month"]]
    for i, a in enumerate(allocations, 1):
        student = a.student
        room = a.room
        block = room.block if room else None
        data.append([i,
                     f"{student.first_name} {student.last_name}" if student else "—",
                     student.admission_number if student else "—",
                     block.name if block else "—",
                     room.name if room else "—",
                     str(a.check_in_date), f"KES {a.monthly_fee or 0:,.0f}"])
    elems.append(pdf_table(data, [1*cm, 5*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm]))
    doc.build(elems)
    return pdf_response(buf, "hostel_allocations.pdf")


# ══════════════════════════════════════════════════════════════════════════════
# TRANSPORT
# ══════════════════════════════════════════════════════════════════════════════

def transport_manifest_excel(db: Session):
    routes = db.query(TransportRoute).filter_by(active=True).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transport Manifest"
    headers = ["Route", "Vehicle", "Reg No", "Driver", "Driver Phone",
               "Stop", "Pickup Time", "Student", "Adm No", "Transport Type"]
    start = add_title_row(ws, "Transport Route Manifest", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for route in routes:
        vehicle = route.vehicle
        stops = db.query(TransportRouteStop).filter_by(route_id=route.id)\
            .order_by(TransportRouteStop.sequence).all()
        for stop in stops:
            assignments = db.query(StudentTransport).filter_by(
                route_id=route.id, stop_id=stop.id, active=True).all()
            if not assignments:
                ws.append([route.name,
                           vehicle.name if vehicle else "—",
                           vehicle.registration_number if vehicle else "—",
                           vehicle.driver_name if vehicle else "—",
                           vehicle.driver_phone if vehicle else "—",
                           stop.name, stop.pickup_time or "—",
                           "— (no students)", "", ""])
            for a in assignments:
                student = a.student
                ws.append([route.name,
                           vehicle.name if vehicle else "—",
                           vehicle.registration_number if vehicle else "—",
                           vehicle.driver_name if vehicle else "—",
                           vehicle.driver_phone if vehicle else "—",
                           stop.name, stop.pickup_time or "—",
                           f"{student.first_name} {student.last_name}" if student else "—",
                           student.admission_number if student else "—",
                           a.transport_type])
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "transport_manifest.xlsx")


def transport_manifest_pdf(db: Session):
    routes = db.query(TransportRoute).filter_by(active=True).all()
    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    total_students = db.query(StudentTransport).filter_by(active=True).count()
    elems = pdf_header(styles, "Transport Route Manifest",
                       f"Routes: {len(routes)}  |  Students: {total_students}  |  "
                       f"Generated: {date.today()}")
    data = [["Route", "Vehicle", "Stop", "Pickup", "Student", "Adm No", "Type"]]
    for route in routes:
        vehicle = route.vehicle
        stops = db.query(TransportRouteStop).filter_by(route_id=route.id)\
            .order_by(TransportRouteStop.sequence).all()
        for stop in stops:
            assignments = db.query(StudentTransport).filter_by(
                route_id=route.id, stop_id=stop.id, active=True).all()
            for a in assignments:
                student = a.student
                data.append([route.name,
                             vehicle.name if vehicle else "—",
                             stop.name, stop.pickup_time or "—",
                             f"{student.first_name} {student.last_name}" if student else "—",
                             student.admission_number if student else "—",
                             a.transport_type])
    elems.append(pdf_table(data, [4*cm, 3*cm, 4*cm, 2.5*cm, 5*cm, 3*cm, 2.5*cm]))
    doc.build(elems)
    return pdf_response(buf, "transport_manifest.pdf")


# ══════════════════════════════════════════════════════════════════════════════
# HEALTH
# ══════════════════════════════════════════════════════════════════════════════

def clinic_visits_excel(db: Session):
    visits = db.query(ClinicVisit).order_by(ClinicVisit.visit_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clinic Visits"
    headers = ["#", "Date", "Student", "Adm No", "Visit Type", "Complaint",
               "Diagnosis", "Treatment", "Disposition", "Attended By"]
    start = add_title_row(ws, "Clinic Visit Records", len(headers),
                          f"Total visits: {len(visits)}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    disp_fills = {
        "referred_hospital": PatternFill("solid", fgColor="FFCCCC"),
        "sent_home":         PatternFill("solid", fgColor="FFF2CC"),
    }
    for i, v in enumerate(visits, 1):
        student = v.student
        ws.append([i, str(v.visit_date),
                   f"{student.first_name} {student.last_name}" if student else "—",
                   student.admission_number if student else "—",
                   str(v.visit_type), v.complaint or "—",
                   v.diagnosis or "—", v.treatment or "—",
                   str(v.disposition), v.attended_by or "—"])
        fill = disp_fills.get(str(v.disposition))
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "clinic_visits.xlsx")


def clinic_visits_pdf(db: Session):
    visits = db.query(ClinicVisit).order_by(ClinicVisit.visit_date.desc()).all()
    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, "Clinic Visit Records",
                       f"Total: {len(visits)} visits  |  Generated: {date.today()}")
    data = [["Date", "Student", "Adm No", "Type", "Complaint", "Disposition"]]
    for v in visits:
        student = v.student
        data.append([str(v.visit_date),
                     f"{student.first_name} {student.last_name}" if student else "—",
                     student.admission_number if student else "—",
                     str(v.visit_type), (v.complaint or "—")[:30],
                     str(v.disposition)])
    elems.append(pdf_table(data, [3*cm, 5*cm, 3*cm, 3*cm, 6*cm, 4*cm]))
    doc.build(elems)
    return pdf_response(buf, "clinic_visits.pdf")


def vaccination_report_excel(db: Session):
    records = db.query(VaccinationRecord).order_by(VaccinationRecord.date_given.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vaccinations"
    headers = ["#", "Student", "Adm No", "Vaccine", "Dose #",
               "Date Given", "Batch No", "Administered By", "Next Dose"]
    start = add_title_row(ws, "Vaccination Records", len(headers),
                          f"Total records: {len(records)}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for i, r in enumerate(records, 1):
        student = r.student
        ws.append([i,
                   f"{student.first_name} {student.last_name}" if student else "—",
                   student.admission_number if student else "—",
                   r.vaccination.name if r.vaccination else "—",
                   r.dose_number, str(r.date_given),
                   r.batch_number or "—", r.administered_by or "—",
                   str(r.next_dose_date) if r.next_dose_date else "—"])
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "vaccinations.xlsx")
