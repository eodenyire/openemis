"""Staff reports — teacher list, TPAD appraisals, leave summary, payroll."""
import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.reports.helpers import (
    style_header_row, style_data_rows, auto_width, add_title_row,
    excel_response, pdf_response, pdf_table, pdf_doc, pdf_header, get_styles,
    NAVY, RED, AMBER, GREEN,
)
from reportlab.lib import colors
from reportlab.platypus import Spacer, Paragraph
from reportlab.lib.units import cm

from app.models.people import Teacher
from app.models.hr import (
    StaffProfile, LeaveRequest, LeaveBalance, LeaveType, LeaveStatus,
    TPADAppraisal, TPADRating, Payslip, PayrollRun, PayrollStatus,
)
from app.models.core import AcademicYear


def _year(db, academic_year_id):
    return (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())


# ── Teacher Directory ─────────────────────────────────────────────────────────

def teacher_list_excel(db: Session):
    teachers = db.query(Teacher).filter_by(active=True).order_by(Teacher.last_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teachers"
    headers = ["#", "Employee ID", "First Name", "Last Name", "Gender",
               "Email", "Phone", "Qualification", "Specialization",
               "Experience (yrs)", "Subjects", "Department"]
    start = add_title_row(ws, "Teacher Directory", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for i, t in enumerate(teachers, 1):
        subjects = ", ".join(s.name for s in t.subjects[:3]) if t.subjects else "—"
        dept = t.main_department.name if t.main_department else "—"
        ws.append([i, t.employee_id, t.first_name, t.last_name,
                   str(t.gender) if t.gender else "—",
                   t.email or "—", t.mobile or t.phone or "—",
                   t.qualification or "—", t.specialization or "—",
                   t.experience_years or 0, subjects, dept])
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "teacher_directory.xlsx")


def teacher_list_pdf(db: Session):
    teachers = db.query(Teacher).filter_by(active=True).order_by(Teacher.last_name).all()
    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, "Teacher Directory",
                       f"Total: {len(teachers)} teachers  |  Generated: {date.today()}")
    data = [["#", "Employee ID", "Name", "Gender", "Qualification", "Specialization", "Subjects"]]
    for i, t in enumerate(teachers, 1):
        subjects = ", ".join(s.name for s in t.subjects[:2]) if t.subjects else "—"
        data.append([i, t.employee_id, f"{t.first_name} {t.last_name}",
                     str(t.gender) if t.gender else "—",
                     t.qualification or "—", t.specialization or "—", subjects])
    elems.append(pdf_table(data, [1*cm, 3*cm, 5*cm, 2*cm, 4*cm, 4*cm, 4*cm]))
    doc.build(elems)
    return pdf_response(buf, "teacher_directory.pdf")


# ── Staff Profiles ────────────────────────────────────────────────────────────

def staff_profiles_excel(db: Session):
    profiles = db.query(StaffProfile).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Profiles"
    headers = ["#", "TSC No", "Job Title", "Employment Type", "Status",
               "Hire Date", "Basic Salary (KES)", "Bank", "Department", "National ID"]
    start = add_title_row(ws, "Staff HR Profiles", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    status_fills = {
        "active":     PatternFill("solid", fgColor="CCFFCC"),
        "on_leave":   PatternFill("solid", fgColor="FFF2CC"),
        "suspended":  PatternFill("solid", fgColor="FFCCCC"),
        "terminated": PatternFill("solid", fgColor="FF9999"),
    }
    for i, p in enumerate(profiles, 1):
        dept = p.department.name if p.department else "—"
        ws.append([i, p.tsc_number or "—", p.job_title or "—",
                   str(p.employment_type) if p.employment_type else "—",
                   str(p.employment_status) if p.employment_status else "—",
                   str(p.hire_date) if p.hire_date else "—",
                   float(p.basic_salary or 0),
                   p.bank_name or "—", dept, p.national_id or "—"])
        fill = status_fills.get(str(p.employment_status), None)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "staff_profiles.xlsx")


# ── TPAD Appraisal Report ─────────────────────────────────────────────────────

def tpad_report_excel(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    q = db.query(TPADAppraisal)
    if year:
        q = q.filter_by(academic_year_id=year.id)
    appraisals = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TPAD Appraisals"
    headers = ["Staff", "TSC No", "Period", "Prof Knowledge", "Lesson Planning",
               "Class Mgmt", "Teaching Method", "Assessment", "Prof Dev",
               "Co-Curricular", "Community", "Avg Score", "Rating", "Submitted"]
    start = add_title_row(ws, f"TPAD Appraisal Report — {year.name if year else 'All'}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    rating_fills = {
        "outstanding":    PatternFill("solid", fgColor="CCFFCC"),
        "exceeds":        PatternFill("solid", fgColor="D6E4F0"),
        "meets":          PatternFill("solid", fgColor="FFF2CC"),
        "below":          PatternFill("solid", fgColor="FFCCCC"),
        "unsatisfactory": PatternFill("solid", fgColor="FF9999"),
    }
    for a in appraisals:
        staff_name = f"{a.staff.teacher.first_name} {a.staff.teacher.last_name}" \
            if a.staff and a.staff.teacher else f"Staff #{a.staff_id}"
        tsc = a.staff.tsc_number if a.staff else "—"
        ws.append([staff_name, tsc, a.appraisal_period,
                   a.professional_knowledge, a.lesson_planning, a.classroom_management,
                   a.teaching_methodology, a.student_assessment, a.professional_development,
                   a.co_curricular, a.community_engagement,
                   a.average_score, str(a.rating) if a.rating else "—",
                   "Yes" if a.is_submitted else "No"])
        fill = rating_fills.get(str(a.rating), None)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"tpad_{year.name if year else 'all'}.xlsx")


def tpad_report_pdf(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    q = db.query(TPADAppraisal)
    if year:
        q = q.filter_by(academic_year_id=year.id)
    appraisals = q.all()

    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, f"TPAD Appraisal Report — {year.name if year else 'All'}",
                       f"Total: {len(appraisals)} appraisals  |  Generated: {date.today()}")
    data = [["Staff", "TSC No", "Period", "Avg Score", "Rating", "Submitted"]]
    for a in appraisals:
        staff_name = f"{a.staff.teacher.first_name} {a.staff.teacher.last_name}" \
            if a.staff and a.staff.teacher else f"Staff #{a.staff_id}"
        data.append([staff_name, a.staff.tsc_number if a.staff else "—",
                     a.appraisal_period, a.average_score,
                     str(a.rating) if a.rating else "—",
                     "Yes" if a.is_submitted else "No"])
    elems.append(pdf_table(data, [6*cm, 3*cm, 3*cm, 3*cm, 4*cm, 2.5*cm]))
    doc.build(elems)
    return pdf_response(buf, f"tpad_{year.name if year else 'all'}.pdf")


# ── Leave Report ──────────────────────────────────────────────────────────────

def leave_report_excel(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    requests = db.query(LeaveRequest).order_by(LeaveRequest.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leave Requests"
    headers = ["Staff", "TSC No", "Leave Type", "Start Date", "End Date",
               "Days", "Status", "Approved By", "Reason"]
    start = add_title_row(ws, "Staff Leave Report", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    status_fills = {
        "approved":  PatternFill("solid", fgColor="CCFFCC"),
        "pending":   PatternFill("solid", fgColor="FFF2CC"),
        "rejected":  PatternFill("solid", fgColor="FFCCCC"),
        "cancelled": PatternFill("solid", fgColor="EEEEEE"),
    }
    for r in requests:
        staff_name = f"{r.staff.teacher.first_name} {r.staff.teacher.last_name}" \
            if r.staff and r.staff.teacher else f"Staff #{r.staff_id}"
        approver = f"{r.approved_by.teacher.first_name} {r.approved_by.teacher.last_name}" \
            if r.approved_by and r.approved_by.teacher else "—"
        ws.append([staff_name, r.staff.tsc_number if r.staff else "—",
                   str(r.leave_type), str(r.start_date), str(r.end_date),
                   r.days_requested, str(r.status), approver, r.reason or "—"])
        fill = status_fills.get(str(r.status), None)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "leave_report.xlsx")


# ── Payroll Report ────────────────────────────────────────────────────────────

def payroll_report_excel(db: Session, payroll_run_id: Optional[int] = None):
    run = (db.query(PayrollRun).filter_by(id=payroll_run_id).first()
           if payroll_run_id
           else db.query(PayrollRun).order_by(PayrollRun.id.desc()).first())
    if not run:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"
    headers = ["#", "Staff", "TSC No", "Job Title", "Basic (KES)", "Gross (KES)",
               "PAYE", "NHIF", "NSSF", "Housing Levy", "Total Deductions", "Net (KES)"]
    start = add_title_row(ws, f"Payroll Report — {run.name}", len(headers),
                          f"Status: {run.status}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for i, slip in enumerate(run.payslips, 1):
        staff_name = f"{slip.staff.teacher.first_name} {slip.staff.teacher.last_name}" \
            if slip.staff and slip.staff.teacher else f"Staff #{slip.staff_id}"
        ws.append([i, staff_name,
                   slip.staff.tsc_number if slip.staff else "—",
                   slip.staff.job_title if slip.staff else "—",
                   float(slip.basic_salary or 0), float(slip.gross_salary or 0),
                   float(slip.paye or 0), float(slip.nhif or 0),
                   float(slip.nssf or 0), float(slip.housing_levy or 0),
                   float(slip.total_deductions or 0), float(slip.net_salary or 0)])

    # Totals
    ws.append(["", "TOTALS", "", "",
               "", float(run.total_gross or 0),
               "", "", "", "",
               float(run.total_deductions or 0), float(run.total_net or 0)])
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True, color="FFFFFF")

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"payroll_{run.name.replace(' ', '_')}.xlsx")


def payroll_report_pdf(db: Session, payroll_run_id: Optional[int] = None):
    run = (db.query(PayrollRun).filter_by(id=payroll_run_id).first()
           if payroll_run_id
           else db.query(PayrollRun).order_by(PayrollRun.id.desc()).first())
    if not run:
        return None

    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, f"Payroll Report — {run.name}",
                       f"Gross: KES {float(run.total_gross or 0):,.2f}  |  "
                       f"Net: KES {float(run.total_net or 0):,.2f}  |  "
                       f"Generated: {date.today()}")
    data = [["#", "Staff", "TSC No", "Basic", "Gross", "PAYE", "NHIF", "NSSF", "Net"]]
    for i, slip in enumerate(run.payslips, 1):
        staff_name = f"{slip.staff.teacher.first_name} {slip.staff.teacher.last_name}" \
            if slip.staff and slip.staff.teacher else f"Staff #{slip.staff_id}"
        data.append([i, staff_name,
                     slip.staff.tsc_number if slip.staff else "—",
                     f"{float(slip.basic_salary or 0):,.0f}",
                     f"{float(slip.gross_salary or 0):,.0f}",
                     f"{float(slip.paye or 0):,.0f}",
                     f"{float(slip.nhif or 0):,.0f}",
                     f"{float(slip.nssf or 0):,.0f}",
                     f"{float(slip.net_salary or 0):,.0f}"])
    elems.append(pdf_table(data, [1*cm, 5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm, 2*cm, 2*cm, 2.5*cm]))
    doc.build(elems)
    return pdf_response(buf, f"payroll_{run.name.replace(' ', '_')}.pdf")
