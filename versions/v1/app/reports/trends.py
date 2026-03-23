"""
Trend & analytics reports — time-series data across all modules.
Covers: attendance trends, fee collection trends, performance progression,
        CBC level distribution over time, exam score trends, LMS engagement,
        health visit trends, library usage trends, staff leave trends.
"""
import io
from datetime import date
from typing import Optional
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, LineChart, Reference
from sqlalchemy import func, extract, case
from sqlalchemy.orm import Session

from app.reports.helpers import (
    style_header_row, style_data_rows, auto_width, add_title_row,
    excel_response, pdf_response, pdf_table, pdf_doc, pdf_header, get_styles,
    NAVY, BLUE, GREEN, AMBER, RED,
)
from reportlab.platypus import Spacer, Paragraph
from reportlab.lib.units import cm
from reportlab.lib import colors

from app.models.attendance import AttendanceLine, AttendanceSheet, AttendanceRegister
from app.models.fees import StudentFeeInvoice, FeePayment
from app.models.exam import Exam, ExamAttendees, ExamSession
from app.models.cbc import ReportCard, ReportCardLine, CBCAssessment, CompetencyScore
from app.models.assignment import Assignment, AssignmentSubmission
from app.models.lms import ClassAssignmentSubmission, QuizAttempt, VirtualClassEnrollment
from app.models.health import ClinicVisit
from app.models.library import MediaMovement
from app.models.hr import LeaveRequest
from app.models.core import AcademicYear, AcademicTerm, Course
from app.models.people import Student, Teacher


# ── Attendance Trend (monthly) ─────────────────────────────────────────────────

def attendance_trend_excel(db: Session, academic_year_id: Optional[int] = None):
    """Monthly attendance rate trend for the whole school."""
    year = (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())

    rows = (db.query(
                extract("month", AttendanceSheet.attendance_date).label("month"),
                extract("year",  AttendanceSheet.attendance_date).label("yr"),
                func.count(AttendanceLine.id).label("total"),
                func.sum(case((AttendanceLine.status == "present", 1), else_=0)).label("present"),
            )
            .join(AttendanceSheet, AttendanceSheet.id == AttendanceLine.sheet_id)
            .join(AttendanceRegister, AttendanceRegister.id == AttendanceSheet.register_id)
            .filter(AttendanceRegister.academic_year_id == year.id if year else True)
            .group_by("yr", "month")
            .order_by("yr", "month")
            .all())

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Trend"
    headers = ["Year", "Month", "Total Sessions", "Present", "Absent", "Rate %"]
    start = add_title_row(ws, f"Monthly Attendance Trend — {year.name if year else 'All'}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    data_start = start + 1
    for r in rows:
        total = r.total or 1
        present = r.present or 0
        rate = round(present / total * 100, 1)
        ws.append([int(r.yr), MONTHS[int(r.month) - 1], total, present,
                   total - present, rate])
        if rate < 75:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")

    style_data_rows(ws, data_start, len(headers))

    # Add line chart
    if rows:
        chart = LineChart()
        chart.title = "Monthly Attendance Rate %"
        chart.style = 10
        chart.y_axis.title = "Rate %"
        chart.x_axis.title = "Month"
        data_ref = Reference(ws, min_col=6, min_row=start, max_row=ws.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        ws.add_chart(chart, f"H{data_start}")

    auto_width(ws)
    return excel_response(wb, f"attendance_trend_{year.name if year else 'all'}.xlsx")


def attendance_trend_pdf(db: Session, academic_year_id: Optional[int] = None):
    year = (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())

    rows = (db.query(
                extract("month", AttendanceSheet.attendance_date).label("month"),
                extract("year",  AttendanceSheet.attendance_date).label("yr"),
                func.count(AttendanceLine.id).label("total"),
                func.sum(case((AttendanceLine.status == "present", 1), else_=0)).label("present"),
            )
            .join(AttendanceSheet, AttendanceSheet.id == AttendanceLine.sheet_id)
            .join(AttendanceRegister, AttendanceRegister.id == AttendanceSheet.register_id)
            .filter(AttendanceRegister.academic_year_id == year.id if year else True)
            .group_by("yr", "month").order_by("yr", "month").all())

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, f"Monthly Attendance Trend — {year.name if year else 'All'}",
                       f"Generated: {date.today()}")
    data = [["Year", "Month", "Total", "Present", "Absent", "Rate %"]]
    for r in rows:
        total = r.total or 1
        present = r.present or 0
        rate = round(present / total * 100, 1)
        data.append([int(r.yr), MONTHS[int(r.month)-1], total, present, total-present, f"{rate}%"])
    elems.append(pdf_table(data, [2*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm]))
    doc.build(elems)
    return pdf_response(buf, f"attendance_trend_{year.name if year else 'all'}.pdf")


# ── Fee Collection Trend (per term) ───────────────────────────────────────────

def fee_collection_trend_excel(db: Session):
    """Fee collection rate per academic term across all years."""
    rows = (db.query(
                AcademicYear.name.label("year"),
                AcademicTerm.name.label("term"),
                func.count(StudentFeeInvoice.id).label("invoices"),
                func.sum(StudentFeeInvoice.total_amount).label("invoiced"),
                func.sum(StudentFeeInvoice.paid_amount).label("paid"),
            )
            .join(AcademicYear, AcademicYear.id == StudentFeeInvoice.academic_year_id)
            .join(AcademicTerm, AcademicTerm.id == StudentFeeInvoice.academic_term_id, isouter=True)
            .group_by(AcademicYear.name, AcademicTerm.name)
            .order_by(AcademicYear.name, AcademicTerm.name)
            .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee Trend"
    headers = ["Academic Year", "Term", "Invoices", "Invoiced (KES)", "Collected (KES)",
               "Outstanding (KES)", "Collection Rate %"]
    start = add_title_row(ws, "Fee Collection Trend — By Term", len(headers),
                          f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for r in rows:
        inv = float(r.invoiced or 0)
        paid = float(r.paid or 0)
        rate = round(paid / inv * 100, 1) if inv else 0
        ws.append([r.year, r.term or "—", r.invoices,
                   round(inv, 2), round(paid, 2), round(inv - paid, 2), rate])
        if rate < 60:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")
        elif rate >= 90:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="CCFFCC")

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "fee_collection_trend.xlsx")


def fee_collection_trend_pdf(db: Session):
    rows = (db.query(
                AcademicYear.name.label("year"),
                AcademicTerm.name.label("term"),
                func.sum(StudentFeeInvoice.total_amount).label("invoiced"),
                func.sum(StudentFeeInvoice.paid_amount).label("paid"),
            )
            .join(AcademicYear, AcademicYear.id == StudentFeeInvoice.academic_year_id)
            .join(AcademicTerm, AcademicTerm.id == StudentFeeInvoice.academic_term_id, isouter=True)
            .group_by(AcademicYear.name, AcademicTerm.name)
            .order_by(AcademicYear.name, AcademicTerm.name)
            .all())

    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf)
    elems = pdf_header(styles, "Fee Collection Trend — By Term", f"Generated: {date.today()}")
    data = [["Year", "Term", "Invoiced (KES)", "Collected (KES)", "Outstanding (KES)", "Rate %"]]
    for r in rows:
        inv = float(r.invoiced or 0)
        paid = float(r.paid or 0)
        data.append([r.year, r.term or "—",
                     f"{inv:,.0f}", f"{paid:,.0f}", f"{inv-paid:,.0f}",
                     f"{round(paid/inv*100,1) if inv else 0}%"])
    elems.append(pdf_table(data, [3.5*cm, 3.5*cm, 4*cm, 4*cm, 4*cm, 2.5*cm]))
    doc.build(elems)
    return pdf_response(buf, "fee_collection_trend.pdf")


# ── Student Performance Progression (per student across terms) ────────────────

def student_performance_progression_excel(db: Session, student_id: int):
    """Track one student's CBC performance level across all terms."""
    student = db.query(Student).get(student_id)
    if not student:
        return None

    cards = (db.query(ReportCard)
             .filter_by(student_id=student_id)
             .order_by(ReportCard.academic_year_id, ReportCard.academic_term_id)
             .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Progression"
    headers = ["Year", "Term", "Grade", "Learning Area", "Performance Level",
               "Teacher Comment", "Days Present", "Days Absent"]
    start = add_title_row(ws,
        f"Performance Progression — {student.first_name} {student.last_name}",
        len(headers),
        f"Adm No: {student.admission_number}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    level_fills = {
        "EE": PatternFill("solid", fgColor="CCFFCC"),
        "ME": PatternFill("solid", fgColor="D6E4F0"),
        "AE": PatternFill("solid", fgColor="FFF2CC"),
        "BE": PatternFill("solid", fgColor="FFCCCC"),
    }
    for card in cards:
        year_name = card.academic_year.name if card.academic_year_id else "—"
        term_name = card.academic_term.name if card.academic_term_id else "—"
        grade = db.query(__import__("app.models.cbc", fromlist=["CBCGradeLevel"])
                         .CBCGradeLevel).get(card.grade_level_id)
        for line in card.lines:
            la = line.learning_area.name if line.learning_area else "—"
            level = str(line.performance_level)
            ws.append([year_name, term_name, grade.name if grade else "—",
                       la, level, line.teacher_comment or "—",
                       card.days_present, card.days_absent])
            fill = level_fills.get(level)
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = fill

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb,
        f"progression_{student.admission_number}.xlsx")


# ── CBC Level Distribution Trend (school-wide per term) ──────────────────────

def cbc_level_distribution_trend_excel(db: Session, academic_year_id: Optional[int] = None):
    """EE/ME/AE/BE counts per term — shows school-wide performance shift."""
    year = (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())

    rows = (db.query(
                AcademicTerm.name.label("term"),
                ReportCardLine.performance_level,
                func.count(ReportCardLine.id).label("count"),
            )
            .join(ReportCard, ReportCard.id == ReportCardLine.report_card_id)
            .join(AcademicTerm, AcademicTerm.id == ReportCard.academic_term_id, isouter=True)
            .filter(ReportCard.academic_year_id == year.id if year else True)
            .group_by(AcademicTerm.name, ReportCardLine.performance_level)
            .order_by(AcademicTerm.name)
            .all())

    # Pivot: term → {EE, ME, AE, BE}
    pivot = defaultdict(lambda: {"EE": 0, "ME": 0, "AE": 0, "BE": 0})
    for r in rows:
        pivot[r.term or "Unknown"][str(r.performance_level)] += r.count

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CBC Level Trend"
    headers = ["Term", "EE", "ME", "AE", "BE", "Total", "EE %", "BE %"]
    start = add_title_row(ws, f"CBC Performance Level Distribution — {year.name if year else 'All'}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for term, dist in pivot.items():
        total = sum(dist.values()) or 1
        ws.append([term, dist["EE"], dist["ME"], dist["AE"], dist["BE"],
                   total,
                   round(dist["EE"] / total * 100, 1),
                   round(dist["BE"] / total * 100, 1)])

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"cbc_level_trend_{year.name if year else 'all'}.xlsx")


# ── Exam Score Trend (per subject across sessions) ────────────────────────────

def exam_score_trend_excel(db: Session, academic_year_id: Optional[int] = None):
    """Average exam score per subject per session — shows subject performance over time."""
    year = (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())

    rows = (db.query(
                ExamSession.name.label("session"),
                Exam.subject_id,
                func.avg(ExamAttendees.marks).label("avg_marks"),
                func.max(ExamAttendees.marks).label("max_marks"),
                func.min(ExamAttendees.marks).label("min_marks"),
                func.count(ExamAttendees.id).label("students"),
                Exam.total_marks,
            )
            .join(Exam, Exam.id == ExamAttendees.exam_id)
            .join(ExamSession, ExamSession.id == Exam.session_id, isouter=True)
            .filter(ExamSession.academic_year_id == year.id if year else True)
            .group_by(ExamSession.name, Exam.subject_id, Exam.total_marks)
            .order_by(ExamSession.name)
            .all())

    from app.models.core import Subject
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Score Trend"
    headers = ["Session", "Subject", "Students", "Avg Marks", "Max", "Min",
               "Total Marks", "Avg %", "Pass Rate %"]
    start = add_title_row(ws, f"Exam Score Trend — {year.name if year else 'All'}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for r in rows:
        subj = db.query(Subject).get(r.subject_id)
        avg = round(float(r.avg_marks or 0), 1)
        total = r.total_marks or 100
        avg_pct = round(avg / total * 100, 1)
        # pass rate: marks >= min_marks
        pass_count = (db.query(ExamAttendees)
                      .join(Exam, Exam.id == ExamAttendees.exam_id)
                      .filter(Exam.subject_id == r.subject_id,
                              ExamAttendees.marks >= Exam.min_marks)
                      .count())
        pass_rate = round(pass_count / (r.students or 1) * 100, 1)
        ws.append([r.session or "—", subj.name if subj else f"#{r.subject_id}",
                   r.students, avg, round(float(r.max_marks or 0), 1),
                   round(float(r.min_marks or 0), 1), total, avg_pct, pass_rate])
        if avg_pct < 50:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")
        elif avg_pct >= 70:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="CCFFCC")

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"exam_score_trend_{year.name if year else 'all'}.xlsx")


# ── LMS Engagement Trend ──────────────────────────────────────────────────────

def lms_engagement_trend_excel(db: Session):
    """Monthly LMS submissions and quiz attempts — digital learning engagement."""
    sub_rows = (db.query(
                    extract("year",  ClassAssignmentSubmission.submitted_at).label("yr"),
                    extract("month", ClassAssignmentSubmission.submitted_at).label("month"),
                    func.count(ClassAssignmentSubmission.id).label("submissions"),
                )
                .filter(ClassAssignmentSubmission.submitted_at.isnot(None))
                .group_by("yr", "month").order_by("yr", "month").all())

    quiz_rows = (db.query(
                    extract("year",  QuizAttempt.submitted_at).label("yr"),
                    extract("month", QuizAttempt.submitted_at).label("month"),
                    func.count(QuizAttempt.id).label("attempts"),
                )
                .filter(QuizAttempt.submitted_at.isnot(None))
                .group_by("yr", "month").order_by("yr", "month").all())

    # Merge by (yr, month)
    merged = defaultdict(lambda: {"submissions": 0, "attempts": 0})
    for r in sub_rows:
        merged[(int(r.yr), int(r.month))]["submissions"] = r.submissions
    for r in quiz_rows:
        merged[(int(r.yr), int(r.month))]["attempts"] = r.attempts

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LMS Engagement"
    headers = ["Year", "Month", "Assignment Submissions", "Quiz Attempts", "Total Activity"]
    start = add_title_row(ws, "LMS Engagement Trend", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for (yr, mo), vals in sorted(merged.items()):
        total = vals["submissions"] + vals["attempts"]
        ws.append([yr, MONTHS[mo - 1], vals["submissions"], vals["attempts"], total])

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "lms_engagement_trend.xlsx")


# ── Health Visit Trend ────────────────────────────────────────────────────────

def health_visit_trend_excel(db: Session):
    """Monthly clinic visit counts by visit type."""
    rows = (db.query(
                extract("year",  ClinicVisit.visit_date).label("yr"),
                extract("month", ClinicVisit.visit_date).label("month"),
                ClinicVisit.visit_type,
                func.count(ClinicVisit.id).label("count"),
            )
            .group_by("yr", "month", ClinicVisit.visit_type)
            .order_by("yr", "month")
            .all())

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Health Visits"
    headers = ["Year", "Month", "Visit Type", "Count"]
    start = add_title_row(ws, "Monthly Health Visit Trend", len(headers),
                          f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for r in rows:
        ws.append([int(r.yr), MONTHS[int(r.month)-1], str(r.visit_type), r.count])

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "health_visit_trend.xlsx")


# ── Library Usage Trend ───────────────────────────────────────────────────────

def library_usage_trend_excel(db: Session):
    """Monthly book issues and returns."""
    rows = (db.query(
                extract("year",  MediaMovement.issue_date).label("yr"),
                extract("month", MediaMovement.issue_date).label("month"),
                func.count(MediaMovement.id).label("issued"),
                func.sum(case((MediaMovement.return_date.isnot(None), 1), else_=0)).label("returned"),
                func.sum(case((MediaMovement.state == "overdue", 1), else_=0)).label("overdue"),
            )
            .group_by("yr", "month").order_by("yr", "month").all())

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Library Usage"
    headers = ["Year", "Month", "Books Issued", "Books Returned", "Overdue"]
    start = add_title_row(ws, "Monthly Library Usage Trend", len(headers),
                          f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for r in rows:
        ws.append([int(r.yr), MONTHS[int(r.month)-1],
                   r.issued, r.returned or 0, r.overdue or 0])
        if (r.overdue or 0) > (r.issued or 1) * 0.3:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "library_usage_trend.xlsx")


# ── Staff Leave Trend ─────────────────────────────────────────────────────────

def staff_leave_trend_excel(db: Session):
    """Monthly approved leave days by leave type."""
    rows = (db.query(
                extract("year",  LeaveRequest.start_date).label("yr"),
                extract("month", LeaveRequest.start_date).label("month"),
                LeaveRequest.leave_type,
                func.count(LeaveRequest.id).label("requests"),
                func.sum(LeaveRequest.days_requested).label("days"),
            )
            .filter(LeaveRequest.status == "approved")
            .group_by("yr", "month", LeaveRequest.leave_type)
            .order_by("yr", "month")
            .all())

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leave Trend"
    headers = ["Year", "Month", "Leave Type", "Requests", "Total Days"]
    start = add_title_row(ws, "Staff Leave Trend (Approved)", len(headers),
                          f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for r in rows:
        ws.append([int(r.yr), MONTHS[int(r.month)-1],
                   str(r.leave_type), r.requests, r.days or 0])

    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "staff_leave_trend.xlsx")
