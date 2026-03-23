"""
Student Result Slips — comprehensive per-student academic report.
Combines: Exam results, Assignments, CBC Assessments, LMS submissions/quizzes,
          Attendance summary, and progressive assessment trend.

Formats: Excel (detailed) + PDF (printable slip).
"""
import io
from datetime import date
from typing import Optional
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, case
from sqlalchemy.orm import Session

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from app.reports.helpers import (
    excel_response, pdf_response, NAVY, BLUE, LIGHT, GREEN, AMBER, RED, WHITE,
    xl_thin_border, xl_center, xl_header_fill, xl_header_font,
)
from app.models.people import Student, StudentCourse
from app.models.core import AcademicYear, AcademicTerm, Course, Subject
from app.models.exam import Exam, ExamAttendees, ExamSession
from app.models.assignment import Assignment, AssignmentSubmission
from app.models.cbc import (
    CBCAssessment, CompetencyScore, ReportCard, ReportCardLine,
    CBCGradeLevel, PerformanceLevel,
)
from app.models.lms import (
    ClassAssignment, ClassAssignmentSubmission,
    Quiz, QuizAttempt,
)
from app.models.attendance import AttendanceLine, AttendanceSheet, AttendanceRegister


# ── Helpers ───────────────────────────────────────────────────────────────────

LEVEL_WEIGHT = {"EE": 4, "ME": 3, "AE": 2, "BE": 1}
LEVEL_COLORS_XL = {
    "EE": "CCFFCC", "ME": "D6E4F0", "AE": "FFF2CC", "BE": "FFCCCC",
}
GRADE_MAP = [
    (80, "A"), (65, "B"), (50, "C"), (40, "D"), (0, "E"),
]

def _grade(pct: float) -> str:
    for threshold, letter in GRADE_MAP:
        if pct >= threshold:
            return letter
    return "E"

def _att_summary(db, student_id, year_id):
    total = (db.query(AttendanceLine)
             .join(AttendanceSheet, AttendanceSheet.id == AttendanceLine.sheet_id)
             .join(AttendanceRegister, AttendanceRegister.id == AttendanceSheet.register_id)
             .filter(AttendanceLine.student_id == student_id,
                     AttendanceRegister.academic_year_id == year_id).count())
    present = (db.query(AttendanceLine)
               .join(AttendanceSheet, AttendanceSheet.id == AttendanceLine.sheet_id)
               .join(AttendanceRegister, AttendanceRegister.id == AttendanceSheet.register_id)
               .filter(AttendanceLine.student_id == student_id,
                       AttendanceRegister.academic_year_id == year_id,
                       AttendanceLine.status == "present").count())
    return total, present


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("SlipTitle", parent=styles["Title"],
                               fontSize=14, textColor=colors.HexColor(f"#{NAVY}"),
                               spaceAfter=2, alignment=TA_CENTER))
    styles.add(ParagraphStyle("SchoolName", parent=styles["Normal"],
                               fontSize=11, textColor=colors.HexColor(f"#{NAVY}"),
                               spaceAfter=2, alignment=TA_CENTER, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle("SubInfo", parent=styles["Normal"],
                               fontSize=8, textColor=colors.grey,
                               spaceAfter=6, alignment=TA_CENTER))
    styles.add(ParagraphStyle("SectionHead", parent=styles["Normal"],
                               fontSize=9, textColor=colors.white,
                               fontName="Helvetica-Bold", alignment=TA_LEFT))
    styles.add(ParagraphStyle("Cell", parent=styles["Normal"],
                               fontSize=8, spaceAfter=0))
    styles.add(ParagraphStyle("Remark", parent=styles["Normal"],
                               fontSize=8, textColor=colors.HexColor("#333333"),
                               spaceAfter=4, leftIndent=4))
    return styles


def _section_header_table(title: str, page_width: float) -> Table:
    t = Table([[title]], colWidths=[page_width])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{NAVY}")),
        ("TEXTCOLOR",  (0, 0), (-1, -1), colors.white),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ]))
    return t


def _std_table(data, col_widths):
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor(f"#{BLUE}")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 7.5),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{LIGHT}")]),
        ("FONTSIZE",      (0, 1), (-1, -1), 7.5),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
    ]))
    return t


# ── Single Student Result Slip — PDF ─────────────────────────────────────────

def student_result_slip_pdf(db: Session, student_id: int,
                             academic_year_id: Optional[int] = None,
                             academic_term_id: Optional[int] = None):
    """
    Comprehensive printable result slip for one student.
    Sections: Student Info | Exam Results | Assignments | CBC Assessments |
              LMS Quizzes | Attendance Summary | Teacher Remarks
    """
    student = db.query(Student).get(student_id)
    if not student:
        return None

    year = (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())
    term = (db.query(AcademicTerm).filter_by(id=academic_term_id).first()
            if academic_term_id else None)

    sc = db.query(StudentCourse).filter_by(student_id=student_id, active=True).first()
    grade = sc.course.name if sc and sc.course else "—"
    roll = sc.roll_number if sc else "—"

    buf = io.BytesIO()
    styles = _pdf_styles()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    PW = A4[0] - 3*cm   # usable page width
    elems = []

    # ── Header ──
    elems.append(Paragraph("STUDENT RESULT SLIP", styles["SlipTitle"]))
    elems.append(Paragraph(
        f"{year.name if year else ''}"
        + (f" — {term.name}" if term else ""),
        styles["SubInfo"]))
    elems.append(HRFlowable(width="100%", thickness=2,
                             color=colors.HexColor(f"#{NAVY}"), spaceAfter=6))

    # ── Student Info table ──
    info = [
        ["Name:", f"{student.first_name} {student.last_name}",
         "Adm No:", student.admission_number],
        ["Grade:", grade, "Roll No:", roll or "—"],
        ["NEMIS UPI:", student.nemis_upi or "—",
         "Gender:", str(student.gender) if student.gender else "—"],
        ["Generated:", str(date.today()), "", ""],
    ]
    info_t = Table(info, colWidths=[2.5*cm, 6*cm, 2.5*cm, 6*cm])
    info_t.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",  (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE",  (0, 0), (-1, -1), 8),
        ("GRID",      (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("BACKGROUND",(0, 0), (-1, -1), colors.HexColor(f"#{LIGHT}")),
        ("TOPPADDING",(0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elems.append(info_t)
    elems.append(Spacer(1, 0.3*cm))

    # ── 1. Exam Results ──
    exam_q = (db.query(ExamAttendees, Exam)
              .join(Exam, Exam.id == ExamAttendees.exam_id)
              .filter(ExamAttendees.student_id == student_id))
    if year:
        exam_q = exam_q.join(ExamSession, ExamSession.id == Exam.session_id, isouter=True)\
                       .filter(ExamSession.academic_year_id == year.id)
    if term:
        exam_q = exam_q.filter(ExamSession.academic_term_id == term.id)
    exam_rows = exam_q.order_by(Exam.start_time).all()

    if exam_rows:
        elems.append(_section_header_table("EXAMINATION RESULTS", PW))
        data = [["Subject", "Exam", "Marks", "Out Of", "%", "Grade", "Status"]]
        for att, exam in exam_rows:
            pct = round((att.marks or 0) / exam.total_marks * 100, 1) if exam.total_marks else 0
            data.append([
                exam.subject.name if exam.subject else "—",
                exam.name[:30],
                att.marks or 0, exam.total_marks,
                f"{pct}%", _grade(pct), att.state,
            ])
        elems.append(_std_table(data, [3.5*cm, 4*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2*cm]))
        elems.append(Spacer(1, 0.3*cm))

    # ── 2. Assignments ──
    assign_q = (db.query(AssignmentSubmission, Assignment)
                .join(Assignment, Assignment.id == AssignmentSubmission.assignment_id)
                .filter(AssignmentSubmission.student_id == student_id))
    if year:
        assign_q = assign_q.filter(
            func.extract("year", Assignment.issued_date) == func.extract("year", year.start_date)
            if hasattr(year, "start_date") and year.start_date else True)
    assign_rows = assign_q.order_by(Assignment.issued_date).all()

    if assign_rows:
        elems.append(_section_header_table("ASSIGNMENTS", PW))
        data = [["Subject", "Assignment", "Issued", "Due", "Marks", "Out Of", "%", "Status"]]
        for sub, asgn in assign_rows:
            pct = round((sub.marks or 0) / asgn.total_marks * 100, 1) if asgn.total_marks else 0
            data.append([
                asgn.subject.name if asgn.subject else "—",
                asgn.name[:25], str(asgn.issued_date), str(asgn.submission_date),
                sub.marks or "—", asgn.total_marks,
                f"{pct}%" if sub.marks else "—", sub.state,
            ])
        elems.append(_std_table(data, [2.5*cm, 3.5*cm, 2*cm, 2*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2*cm]))
        elems.append(Spacer(1, 0.3*cm))

    # ── 3. CBC Assessments ──
    cbc_q = (db.query(CompetencyScore, CBCAssessment)
             .join(CBCAssessment, CBCAssessment.id == CompetencyScore.assessment_id)
             .filter(CompetencyScore.student_id == student_id))
    if year:
        cbc_q = cbc_q.filter(CBCAssessment.academic_year == year.name)
    if term:
        cbc_q = cbc_q.filter(CBCAssessment.term == term.name)
    cbc_rows = cbc_q.order_by(CBCAssessment.assessment_date).all()

    if cbc_rows:
        elems.append(_section_header_table("CBC COMPETENCY ASSESSMENTS", PW))
        data = [["Learning Area", "Assessment", "Type", "Date", "Level", "Raw Score", "Remarks"]]
        for score, assessment in cbc_rows:
            la = assessment.learning_area.name if assessment.learning_area else "—"
            data.append([
                la, assessment.name[:25], assessment.assessment_type,
                str(assessment.assessment_date) if assessment.assessment_date else "—",
                str(score.performance_level),
                score.raw_score if score.raw_score is not None else "—",
                (score.teacher_remarks or "—")[:30],
            ])
        elems.append(_std_table(data, [3*cm, 3*cm, 2*cm, 2*cm, 1.5*cm, 2*cm, 3.5*cm]))
        elems.append(Spacer(1, 0.3*cm))

    # ── 4. LMS Quizzes ──
    quiz_q = (db.query(QuizAttempt, Quiz)
              .join(Quiz, Quiz.id == QuizAttempt.quiz_id)
              .filter(QuizAttempt.student_id == student_id,
                      QuizAttempt.is_submitted == True))
    quiz_rows = quiz_q.order_by(QuizAttempt.submitted_at).all()

    if quiz_rows:
        elems.append(_section_header_table("ONLINE QUIZZES", PW))
        data = [["Quiz", "Classroom", "Score", "Out Of", "%", "Submitted"]]
        for attempt, quiz in quiz_rows:
            pct = round((attempt.score or 0) / quiz.total_marks * 100, 1) if quiz.total_marks else 0
            classroom = quiz.classroom.name if quiz.classroom else "—"
            data.append([
                quiz.title[:30], classroom,
                attempt.score or 0, quiz.total_marks,
                f"{pct}%",
                str(attempt.submitted_at)[:16] if attempt.submitted_at else "—",
            ])
        elems.append(_std_table(data, [4*cm, 3*cm, 1.5*cm, 1.5*cm, 1.5*cm, 4*cm]))
        elems.append(Spacer(1, 0.3*cm))

    # ── 5. Attendance Summary ──
    if year:
        total_att, present_att = _att_summary(db, student_id, year.id)
        absent = total_att - present_att
        rate = round(present_att / total_att * 100, 1) if total_att else 0
        elems.append(_section_header_table("ATTENDANCE SUMMARY", PW))
        att_data = [
            ["Total Sessions", "Present", "Absent", "Attendance Rate"],
            [total_att, present_att, absent, f"{rate}%"],
        ]
        att_t = Table(att_data, colWidths=[PW/4]*4)
        att_t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor(f"#{BLUE}")),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("BACKGROUND",    (0, 1), (-1, 1),
             colors.HexColor("#CCFFCC") if rate >= 75 else colors.HexColor("#FFCCCC")),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elems.append(att_t)
        elems.append(Spacer(1, 0.3*cm))

    # ── 6. CBC Report Card Remarks ──
    card_q = db.query(ReportCard).filter_by(student_id=student_id)
    if year:
        card_q = card_q.filter_by(academic_year_id=year.id)
    if term:
        card_q = card_q.filter_by(academic_term_id=term.id)
    card = card_q.first()

    if card:
        elems.append(_section_header_table("TEACHER & PRINCIPAL REMARKS", PW))
        remarks_data = [
            ["Class Teacher:", card.teacher_remarks or "—"],
            ["Principal:",     card.principal_remarks or "—"],
        ]
        rem_t = Table(remarks_data, colWidths=[3*cm, PW - 3*cm])
        rem_t.setStyle(TableStyle([
            ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE",  (0, 0), (-1, -1), 8),
            ("GRID",      (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("BACKGROUND",(0, 0), (-1, -1), colors.HexColor(f"#{LIGHT}")),
            ("TOPPADDING",(0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elems.append(rem_t)

    # ── Footer ──
    elems.append(Spacer(1, 0.5*cm))
    elems.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor("#CCCCCC"), spaceAfter=4))
    elems.append(Paragraph(
        "This is a computer-generated result slip. "
        "For queries contact the school administration.",
        ParagraphStyle("Footer", parent=styles["Normal"],
                       fontSize=7, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(elems)
    return pdf_response(buf,
        f"result_slip_{student.admission_number}_{year.name if year else 'all'}.pdf")


# ── Bulk Result Slips — all students in a grade (Excel) ──────────────────────

def bulk_result_slips_excel(db: Session, course_id: Optional[int] = None,
                             academic_year_id: Optional[int] = None,
                             academic_term_id: Optional[int] = None):
    """
    One Excel workbook with one sheet per student — full result slip.
    Useful for printing all slips for a class at once.
    """
    year = (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())
    term = (db.query(AcademicTerm).filter_by(id=academic_term_id).first()
            if academic_term_id else None)

    q = db.query(Student, Course.name.label("grade"))\
        .join(StudentCourse, StudentCourse.student_id == Student.id)\
        .join(Course, Course.id == StudentCourse.course_id)\
        .filter(StudentCourse.active == True, Student.active == True)
    if year:
        q = q.filter(StudentCourse.academic_year_id == year.id)
    if course_id:
        q = q.filter(StudentCourse.course_id == course_id)
    students = q.order_by(Course.name, Student.last_name).all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    thin = xl_thin_border()
    hdr_fill = xl_header_fill()
    hdr_font = xl_header_font()
    center = xl_center()

    for student, grade in students:
        # Sheet name max 31 chars
        sheet_name = f"{student.admission_number}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"RESULT SLIP — {student.first_name} {student.last_name}"
        ws["A1"].font = Font(bold=True, size=13, color=NAVY)
        ws["A1"].alignment = center

        ws.merge_cells("A2:H2")
        ws["A2"] = (f"Adm: {student.admission_number}  |  Grade: {grade}  |  "
                    f"NEMIS: {student.nemis_upi or '—'}  |  "
                    f"Year: {year.name if year else '—'}"
                    + (f"  |  Term: {term.name}" if term else ""))
        ws["A2"].font = Font(italic=True, size=9, color="555555")
        ws["A2"].alignment = center

        row = 4

        def write_section(title, headers, data_rows):
            nonlocal row
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=len(headers))
            cell = ws.cell(row=row, column=1, value=title)
            cell.fill = xl_header_fill(NAVY)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.alignment = center
            row += 1
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.fill = xl_header_fill(BLUE)
                c.font = hdr_font
                c.alignment = center
                c.border = thin
            row += 1
            for dr in data_rows:
                for ci, val in enumerate(dr, 1):
                    c = ws.cell(row=row, column=ci, value=val)
                    c.border = thin
                    if row % 2 == 0:
                        c.fill = PatternFill("solid", fgColor=LIGHT)
                row += 1
            row += 1  # blank separator

        # Exam results
        exam_q = (db.query(ExamAttendees, Exam)
                  .join(Exam, Exam.id == ExamAttendees.exam_id)
                  .filter(ExamAttendees.student_id == student.id))
        if year:
            from app.models.exam import ExamSession as ES
            exam_q = exam_q.join(ES, ES.id == Exam.session_id, isouter=True)\
                           .filter(ES.academic_year_id == year.id)
        exam_rows = exam_q.order_by(Exam.start_time).all()
        if exam_rows:
            write_section("EXAMINATION RESULTS",
                ["Subject", "Exam", "Marks", "Out Of", "%", "Grade", "Status"],
                [[exam.subject.name if exam.subject else "—", exam.name,
                  att.marks or 0, exam.total_marks,
                  round((att.marks or 0) / exam.total_marks * 100, 1) if exam.total_marks else 0,
                  _grade(round((att.marks or 0) / exam.total_marks * 100, 1) if exam.total_marks else 0),
                  att.state]
                 for att, exam in exam_rows])

        # Assignments
        assign_rows = (db.query(AssignmentSubmission, Assignment)
                       .join(Assignment, Assignment.id == AssignmentSubmission.assignment_id)
                       .filter(AssignmentSubmission.student_id == student.id)
                       .order_by(Assignment.issued_date).all())
        if assign_rows:
            write_section("ASSIGNMENTS",
                ["Subject", "Assignment", "Issued", "Due", "Marks", "Out Of", "%", "Status"],
                [[asgn.subject.name if asgn.subject else "—", asgn.name,
                  str(asgn.issued_date), str(asgn.submission_date),
                  sub.marks or "—", asgn.total_marks,
                  round((sub.marks or 0) / asgn.total_marks * 100, 1) if asgn.total_marks and sub.marks else "—",
                  sub.state]
                 for sub, asgn in assign_rows])

        # CBC Assessments
        cbc_q = (db.query(CompetencyScore, CBCAssessment)
                 .join(CBCAssessment, CBCAssessment.id == CompetencyScore.assessment_id)
                 .filter(CompetencyScore.student_id == student.id))
        if year:
            cbc_q = cbc_q.filter(CBCAssessment.academic_year == year.name)
        cbc_rows = cbc_q.order_by(CBCAssessment.assessment_date).all()
        if cbc_rows:
            write_section("CBC ASSESSMENTS",
                ["Learning Area", "Assessment", "Type", "Date", "Level", "Raw Score"],
                [[assessment.learning_area.name if assessment.learning_area else "—",
                  assessment.name, assessment.assessment_type,
                  str(assessment.assessment_date) if assessment.assessment_date else "—",
                  str(score.performance_level), score.raw_score or "—"]
                 for score, assessment in cbc_rows])

        # Attendance
        if year:
            total_att, present_att = _att_summary(db, student.id, year.id)
            rate = round(present_att / total_att * 100, 1) if total_att else 0
            write_section("ATTENDANCE",
                ["Total Sessions", "Present", "Absent", "Rate %"],
                [[total_att, present_att, total_att - present_att, f"{rate}%"]])

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    label = f"_{year.name}" if year else ""
    label += f"_{term.name}" if term else ""
    return excel_response(wb, f"result_slips{label}.xlsx")


# ── Progressive Assessment Summary (per student, all terms) ──────────────────

def progressive_assessment_excel(db: Session, student_id: int):
    """
    Shows a student's marks/levels across every assessment type over time.
    Useful for tracking improvement or decline term-by-term.
    """
    student = db.query(Student).get(student_id)
    if not student:
        return None

    wb = openpyxl.Workbook()

    # Sheet 1: Exam progression
    ws1 = wb.active
    ws1.title = "Exam Progression"
    headers = ["Session", "Subject", "Marks", "Total", "%", "Grade"]
    ws1.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws1.cell(row=1, column=ci)
        c.fill = xl_header_fill()
        c.font = xl_header_font()
        c.alignment = xl_center()
        c.border = xl_thin_border()

    from app.models.exam import ExamSession as ES
    rows = (db.query(ExamAttendees, Exam, ES)
            .join(Exam, Exam.id == ExamAttendees.exam_id)
            .join(ES, ES.id == Exam.session_id, isouter=True)
            .filter(ExamAttendees.student_id == student_id)
            .order_by(ES.start_date, Exam.subject_id).all())
    for att, exam, sess in rows:
        pct = round((att.marks or 0) / exam.total_marks * 100, 1) if exam.total_marks else 0
        ws1.append([sess.name if sess else "—",
                    exam.subject.name if exam.subject else "—",
                    att.marks or 0, exam.total_marks, pct, _grade(pct)])
        if pct < 50:
            for col in range(1, len(headers) + 1):
                ws1.cell(row=ws1.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")

    # Sheet 2: CBC level progression
    ws2 = wb.create_sheet("CBC Progression")
    headers2 = ["Year", "Term", "Learning Area", "Assessment", "Level", "Raw Score"]
    ws2.append(headers2)
    for ci in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=ci)
        c.fill = xl_header_fill()
        c.font = xl_header_font()
        c.alignment = xl_center()
        c.border = xl_thin_border()

    cbc_rows = (db.query(CompetencyScore, CBCAssessment)
                .join(CBCAssessment, CBCAssessment.id == CompetencyScore.assessment_id)
                .filter(CompetencyScore.student_id == student_id)
                .order_by(CBCAssessment.academic_year, CBCAssessment.term,
                          CBCAssessment.assessment_date).all())
    for score, assessment in cbc_rows:
        la = assessment.learning_area.name if assessment.learning_area else "—"
        level = str(score.performance_level)
        ws2.append([assessment.academic_year or "—", assessment.term or "—",
                    la, assessment.name, level, score.raw_score or "—"])
        fill_color = LEVEL_COLORS_XL.get(level)
        if fill_color:
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=ws2.max_row, column=col).fill = PatternFill("solid", fgColor=fill_color)

    # Sheet 3: Assignment progression
    ws3 = wb.create_sheet("Assignment Progression")
    headers3 = ["Subject", "Assignment", "Issued", "Marks", "Total", "%", "Status"]
    ws3.append(headers3)
    for ci in range(1, len(headers3) + 1):
        c = ws3.cell(row=1, column=ci)
        c.fill = xl_header_fill()
        c.font = xl_header_font()
        c.alignment = xl_center()
        c.border = xl_thin_border()

    assign_rows = (db.query(AssignmentSubmission, Assignment)
                   .join(Assignment, Assignment.id == AssignmentSubmission.assignment_id)
                   .filter(AssignmentSubmission.student_id == student_id)
                   .order_by(Assignment.issued_date).all())
    for sub, asgn in assign_rows:
        pct = round((sub.marks or 0) / asgn.total_marks * 100, 1) if asgn.total_marks and sub.marks else 0
        ws3.append([asgn.subject.name if asgn.subject else "—",
                    asgn.name, str(asgn.issued_date),
                    sub.marks or "—", asgn.total_marks,
                    f"{pct}%" if sub.marks else "—", sub.state])

    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    return excel_response(wb, f"progressive_{student.admission_number}.xlsx")
