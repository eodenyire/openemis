"""Academic reports — exams, CBC assessments, LMS assignments, timetable."""
import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.reports.helpers import (
    style_header_row, style_data_rows, auto_width, add_title_row,
    excel_response, pdf_response, pdf_table, pdf_doc, pdf_header, get_styles,
)
from reportlab.platypus import Spacer
from reportlab.lib.units import cm

from app.models.exam import Exam, ExamAttendees, ExamSession
from app.models.cbc import (
    CBCAssessment, CompetencyScore, ReportCard, ReportCardLine,
    CBCGradeLevel, PerformanceLevel,
)
from app.models.lms import (
    VirtualClassroom, ClassAssignment, ClassAssignmentSubmission,
    Quiz, QuizAttempt, LessonPlan, SchemeOfWork,
)
from app.models.core import AcademicYear, Course, Subject
from app.models.people import Student, Teacher


def _year(db, academic_year_id):
    return (db.query(AcademicYear).filter_by(id=academic_year_id).first()
            if academic_year_id
            else db.query(AcademicYear).order_by(AcademicYear.id.desc()).first())


# ── Exam Results ──────────────────────────────────────────────────────────────

def exam_results_excel(db: Session, exam_session_id: Optional[int] = None):
    q = db.query(ExamAttendees, Exam, Student)\
        .join(Exam, Exam.id == ExamAttendees.exam_id)\
        .join(Student, Student.id == ExamAttendees.student_id)
    if exam_session_id:
        q = q.filter(Exam.session_id == exam_session_id)
    rows = q.order_by(Exam.name, Student.last_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results"
    headers = ["Exam", "Subject", "Student", "Adm No", "Marks", "Total Marks",
               "Percentage %", "State", "Grade"]
    start = add_title_row(ws, "Exam Results Report", len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    fills = {
        "pass": PatternFill("solid", fgColor="CCFFCC"),
        "fail": PatternFill("solid", fgColor="FFCCCC"),
        "absent": PatternFill("solid", fgColor="EEEEEE"),
    }
    for attendee, exam, student in rows:
        pct = round((attendee.marks or 0) / exam.total_marks * 100, 1) if exam.total_marks else 0
        # Simple grade
        grade = "A" if pct >= 80 else "B" if pct >= 65 else "C" if pct >= 50 else "D" if pct >= 40 else "E"
        ws.append([exam.name,
                   exam.subject.name if exam.subject else "—",
                   f"{student.first_name} {student.last_name}",
                   student.admission_number,
                   attendee.marks or 0, exam.total_marks, pct,
                   attendee.state, grade])
        fill = fills.get(attendee.state)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "exam_results.xlsx")


def exam_results_pdf(db: Session, exam_session_id: Optional[int] = None):
    q = db.query(ExamAttendees, Exam, Student)\
        .join(Exam, Exam.id == ExamAttendees.exam_id)\
        .join(Student, Student.id == ExamAttendees.student_id)
    if exam_session_id:
        q = q.filter(Exam.session_id == exam_session_id)
    rows = q.order_by(Exam.name, Student.last_name).all()

    buf = io.BytesIO()
    styles = get_styles()
    doc = pdf_doc(buf, landscape_mode=True)
    elems = pdf_header(styles, "Exam Results Report", f"Generated: {date.today()}")
    data = [["Exam", "Subject", "Student", "Adm No", "Marks", "Total", "%", "State"]]
    for attendee, exam, student in rows:
        pct = round((attendee.marks or 0) / exam.total_marks * 100, 1) if exam.total_marks else 0
        data.append([exam.name[:25],
                     exam.subject.name[:15] if exam.subject else "—",
                     f"{student.first_name} {student.last_name}",
                     student.admission_number,
                     attendee.marks or 0, exam.total_marks, f"{pct}%",
                     attendee.state])
    elems.append(pdf_table(data, [4*cm, 3*cm, 5*cm, 3*cm, 2*cm, 2*cm, 2*cm, 2*cm]))
    doc.build(elems)
    return pdf_response(buf, "exam_results.pdf")


# ── CBC Report Cards ──────────────────────────────────────────────────────────

def report_cards_excel(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    q = db.query(ReportCard)
    if year:
        q = q.filter_by(academic_year_id=year.id)
    cards = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Cards"
    headers = ["Student", "Adm No", "Grade", "Term", "Days Present", "Days Absent",
               "Learning Areas", "EE", "ME", "AE", "BE", "Published"]
    start = add_title_row(ws, f"CBC Report Cards — {year.name if year else ''}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for card in cards:
        student = db.query(Student).get(card.student_id)
        grade = db.query(CBCGradeLevel).get(card.grade_level_id)
        term = card.academic_term.name if card.academic_term_id else "—"
        dist = {"EE": 0, "ME": 0, "AE": 0, "BE": 0}
        for line in card.lines:
            dist[str(line.performance_level)] = dist.get(str(line.performance_level), 0) + 1
        ws.append([
            f"{student.first_name} {student.last_name}" if student else "—",
            student.admission_number if student else "—",
            grade.name if grade else "—", term,
            card.days_present, card.days_absent,
            len(card.lines),
            dist["EE"], dist["ME"], dist["AE"], dist["BE"],
            "Yes" if card.is_published else "No",
        ])
        if not card.is_published:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"report_cards_{year.name if year else 'all'}.xlsx")


# ── LMS Assignment Submissions ────────────────────────────────────────────────

def lms_submissions_excel(db: Session):
    submissions = db.query(ClassAssignmentSubmission)\
        .order_by(ClassAssignmentSubmission.submitted_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LMS Submissions"
    headers = ["Assignment", "Classroom", "Student", "Adm No",
               "Submitted At", "Status", "Marks", "Total Marks", "Feedback"]
    start = add_title_row(ws, "LMS Assignment Submissions", len(headers),
                          f"Total: {len(submissions)}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    fills = {
        "graded":    PatternFill("solid", fgColor="CCFFCC"),
        "late":      PatternFill("solid", fgColor="FFCCCC"),
        "pending":   PatternFill("solid", fgColor="FFF2CC"),
    }
    for sub in submissions:
        assignment = sub.assignment
        classroom = assignment.classroom if assignment else None
        student = db.query(Student).get(sub.student_id)
        ws.append([
            assignment.title if assignment else "—",
            classroom.name if classroom else "—",
            f"{student.first_name} {student.last_name}" if student else "—",
            student.admission_number if student else "—",
            str(sub.submitted_at)[:16] if sub.submitted_at else "—",
            str(sub.status),
            sub.marks_obtained or "—",
            assignment.total_marks if assignment else "—",
            (sub.feedback or "—")[:50],
        ])
        fill = fills.get(str(sub.status))
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "lms_submissions.xlsx")


# ── Quiz Attempts ─────────────────────────────────────────────────────────────

def quiz_attempts_excel(db: Session):
    attempts = db.query(QuizAttempt).filter_by(is_submitted=True)\
        .order_by(QuizAttempt.submitted_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz Attempts"
    headers = ["Quiz", "Classroom", "Student", "Adm No",
               "Score", "Total Marks", "Percentage %", "Submitted At"]
    start = add_title_row(ws, "Quiz Attempt Results", len(headers),
                          f"Total: {len(attempts)}  |  Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    for a in attempts:
        quiz = a.quiz
        classroom = quiz.classroom if quiz else None
        student = db.query(Student).get(a.student_id)
        pct = round((a.score or 0) / quiz.total_marks * 100, 1) if quiz and quiz.total_marks else 0
        ws.append([
            quiz.title if quiz else "—",
            classroom.name if classroom else "—",
            f"{student.first_name} {student.last_name}" if student else "—",
            student.admission_number if student else "—",
            a.score or 0, quiz.total_marks if quiz else "—",
            pct, str(a.submitted_at)[:16] if a.submitted_at else "—",
        ])
        if pct < 50:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor="FFCCCC")
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, "quiz_attempts.xlsx")


# ── Lesson Plans ──────────────────────────────────────────────────────────────

def lesson_plans_excel(db: Session, academic_year_id: Optional[int] = None):
    year = _year(db, academic_year_id)
    q = db.query(LessonPlan, SchemeOfWork)\
        .join(SchemeOfWork, SchemeOfWork.id == LessonPlan.scheme_id)
    if year:
        q = q.filter(SchemeOfWork.academic_year_id == year.id)
    rows = q.order_by(SchemeOfWork.subject_id, LessonPlan.week_number).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lesson Plans"
    headers = ["Teacher", "Subject", "Grade", "Week", "Lesson #",
               "Title", "Duration (min)", "Status"]
    start = add_title_row(ws, f"Lesson Plans — {year.name if year else ''}",
                          len(headers), f"Generated: {date.today()}")
    ws.append(headers)
    style_header_row(ws, start, len(headers))

    status_fills = {
        "approved":  PatternFill("solid", fgColor="CCFFCC"),
        "submitted": PatternFill("solid", fgColor="D6E4F0"),
        "draft":     PatternFill("solid", fgColor="FFF2CC"),
        "rejected":  PatternFill("solid", fgColor="FFCCCC"),
    }
    for plan, scheme in rows:
        teacher = scheme.teacher
        ws.append([
            f"{teacher.first_name} {teacher.last_name}" if teacher else "—",
            scheme.subject.name if scheme.subject else "—",
            scheme.course.name if scheme.course else "—",
            plan.week_number, plan.lesson_number,
            plan.title, plan.duration_minutes, str(plan.status),
        ])
        fill = status_fills.get(str(plan.status))
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    style_data_rows(ws, start + 1, len(headers))
    auto_width(ws)
    return excel_response(wb, f"lesson_plans_{year.name if year else 'all'}.xlsx")
